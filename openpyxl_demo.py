'''
Created on 15-Sep-2023

@author: bborade
'''
import os
import openpyxl

keys_template = []
raw_data = []
workbook = openpyxl.load_workbook(os.path.dirname(__file__) + '/data_files/UAT_Test_Data.xlsx')
print(dir(workbook))
sheet = workbook.active
print(dir(sheet))
print(sheet.max_row)
print(sheet.max_column)
for data_cell in sheet['2']:
    if data_cell.value != None:
        keys_template.append(str(data_cell.value))
print(keys_template)
print(len(keys_template))
for row in range(3, sheet.max_row):
    if sheet.cell(row, 1).value == "Regis":
        req_data = {}
        for column in range(1, len(keys_template)):
            print(sheet.cell(row, 1).value)
            print(keys_template[column])
            print(sheet.cell(row, column).value)
            if sheet.cell(row, column).value != None:
                req_data[keys_template[column-1]] = str(sheet.cell(row, column).value)
            else:
                req_data[keys_template[column-1]] = ''
            print(req_data)
        raw_data.append(req_data)
print(len(raw_data))
print(raw_data)
'''
workbook
['_Workbook__write_only', '__class__', '__contains__', '__delattr__', '__delitem__', '__dict__', '__dir__', '__doc__', '__eq__', '__format__',
'__ge__', '__getattribute__', '__getitem__', '__gt__', '__hash__', '__init__', '__init_subclass__', '__iter__', '__le__', '__lt__', '__module__',
'__ne__', '__new__', '__reduce__', '__reduce_ex__', '__repr__', '__setattr__', '__sizeof__', '__str__', '__subclasshook__', '__weakref__',
'_active_sheet_index', '_add_sheet', '_alignments', '_borders', '_cell_styles', '_colors', '_data_only', '_date_formats', '_differential_styles',
'_duplicate_name', '_epoch', '_external_links', '_fills', '_fonts', '_named_styles', '_number_formats', '_pivots', '_protections', '_read_only',
'_setup_styles', '_sheets', '_table_styles', '_timedelta_formats', 'active', 'add_named_style', 'calculation', 'chartsheets', 'close', 'code_name',
'copy_worksheet', 'create_chartsheet', 'create_named_range', 'create_sheet', 'custom_doc_props', 'data_only', 'defined_names', 'encoding', 'epoch',
'excel_base_date', 'get_index', 'get_sheet_by_name', 'get_sheet_names', 'index', 'is_template', 'iso_dates', 'loaded_theme', 'mime_type', 'move_sheet',
'named_styles', 'path', 'properties', 'read_only', 'rels', 'remove', 'remove_sheet', 'save', 'security', 'shared_strings', 'sheetnames', 'style_names',
'template', 'vba_archive', 'views', 'worksheets', 'write_only']

sheet
['BREAK_COLUMN', 'BREAK_NONE', 'BREAK_ROW', 'HeaderFooter', 'ORIENTATION_LANDSCAPE', 'ORIENTATION_PORTRAIT', 'PAPERSIZE_A3', 'PAPERSIZE_A4',
'PAPERSIZE_A4_SMALL', 'PAPERSIZE_A5', 'PAPERSIZE_EXECUTIVE', 'PAPERSIZE_LEDGER', 'PAPERSIZE_LEGAL', 'PAPERSIZE_LETTER', 'PAPERSIZE_LETTER_SMALL',
'PAPERSIZE_STATEMENT', 'PAPERSIZE_TABLOID', 'SHEETSTATE_HIDDEN', 'SHEETSTATE_VERYHIDDEN', 'SHEETSTATE_VISIBLE', '_WorkbookChild__title', '__class__',
'__delattr__', '__delitem__', '__dict__', '__dir__', '__doc__', '__eq__', '__format__', '__ge__', '__getattribute__', '__getitem__', '__gt__',
'__hash__', '__init__', '__init_subclass__', '__iter__', '__le__', '__lt__', '__module__', '__ne__', '__new__', '__reduce__', '__reduce_ex__',
'__repr__', '__setattr__', '__setitem__', '__sizeof__', '__str__', '__subclasshook__', '__weakref__', '_add_cell', '_add_column', '_add_row',
'_cells', '_cells_by_col', '_cells_by_row', '_charts', '_clean_merge_range', '_comments', '_current_row', '_default_title', '_drawing', '_get_cell',
'_hyperlinks', '_id', '_images', '_invalid_row', '_move_cell', '_move_cells', '_parent', '_path', '_pivots', '_print_area', '_print_cols',
'_print_rows', '_rel_type', '_rels', '_setup', '_tables', 'active_cell', 'add_chart', 'add_data_validation', 'add_image', 'add_pivot', 'add_table',
'append', 'array_formulae', 'auto_filter', 'calculate_dimension', 'cell', 'col_breaks', 'column_dimensions', 'columns', 'conditional_formatting',
'data_validations', 'defined_names', 'delete_cols', 'delete_rows', 'dimensions', 'encoding', 'evenFooter', 'evenHeader', 'firstFooter', 'firstHeader',
'freeze_panes', 'insert_cols', 'insert_rows', 'iter_cols', 'iter_rows', 'legacy_drawing', 'max_column', 'max_row', 'merge_cells', 'merged_cell_ranges',
'merged_cells', 'mime_type', 'min_column', 'min_row', 'move_range', 'oddFooter', 'oddHeader', 'page_margins', 'page_setup', 'parent', 'path',
'print_area', 'print_options', 'print_title_cols', 'print_title_rows', 'print_titles', 'protection', 'row_breaks', 'row_dimensions', 'rows', 'scenarios',
'selected_cell', 'set_printer_settings', 'sheet_format', 'sheet_properties', 'sheet_state', 'sheet_view', 'show_gridlines', 'tables', 'title',
'unmerge_cells', 'values', 'views']
'''
